#!/usr/bin/env python3
"""
XLS Processor — reads xlsx structure and executes AI-generated openpyxl code.
Supports .xls (legacy) via xlrd → openpyxl conversion.
Called by Express via subprocess.

IMPORTANT: multer saves uploaded files WITHOUT extension (/tmp/<hash>).
openpyxl checks the file extension before opening, so we always create
a temp copy with the correct extension before passing to openpyxl.
"""

import sys
import json
import os
import re
import traceback
import tempfile
import shutil


def get_ext(original_name: str) -> str:
    """Get lowercase extension from original filename, e.g. '.xlsx' or '.xls'."""
    if not original_name:
        return ".xlsx"
    m = re.search(r'\.(xlsx|xlsm|xls|xltx|xltm)$', original_name, re.IGNORECASE)
    return m.group(0).lower() if m else ".xlsx"


def make_temp_copy(path: str, ext: str) -> str:
    """Create a temp copy of the file with the given extension. Caller must delete it."""
    tmp = tempfile.mktemp(suffix=ext, dir=tempfile.gettempdir())
    shutil.copy2(path, tmp)
    return tmp


def convert_xls_to_xlsx(xls_path: str) -> str:
    """
    Convert a legacy .xls file to a temporary .xlsx using xlrd + openpyxl.
    Returns path to the new .xlsx file (caller must delete it when done).
    """
    import xlrd
    import openpyxl
    import datetime

    wb_xls = xlrd.open_workbook(xls_path, formatting_info=False)
    wb_xlsx = openpyxl.Workbook()
    wb_xlsx.remove(wb_xlsx.active)  # remove default empty sheet

    for sheet_idx in range(wb_xls.nsheets):
        ws_xls = wb_xls.sheet_by_index(sheet_idx)
        ws_xlsx = wb_xlsx.create_sheet(title=ws_xls.name)

        for row_idx in range(ws_xls.nrows):
            for col_idx in range(ws_xls.ncols):
                cell = ws_xls.cell(row_idx, col_idx)
                # xlrd type: 0=empty, 1=text, 2=number, 3=date, 4=bool, 5=error
                if cell.ctype == 1:
                    value = cell.value
                elif cell.ctype == 2:
                    value = int(cell.value) if cell.value == int(cell.value) else cell.value
                elif cell.ctype == 4:
                    value = bool(cell.value)
                elif cell.ctype == 3:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell.value, wb_xls.datemode)
                        if date_tuple[3:] == (0, 0, 0):
                            value = datetime.date(*date_tuple[:3])
                        else:
                            value = datetime.datetime(*date_tuple)
                    except Exception:
                        value = cell.value
                else:
                    value = cell.value if cell.ctype != 0 else None

                if value is not None and value != "":
                    ws_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    tmp_path = tempfile.mktemp(suffix=".xlsx", dir=tempfile.gettempdir())
    wb_xlsx.save(tmp_path)
    return tmp_path


def prepare_file(path: str, original_name: str) -> tuple:
    """
    Prepare file for openpyxl:
    - If .xls: convert to .xlsx, return (converted_path, [converted_path])
    - Otherwise: copy with correct extension, return (copy_path, [copy_path])
    temps list contains paths to clean up.
    """
    ext = get_ext(original_name)
    temps = []

    if ext == ".xls":
        # Need a temp copy with .xls extension first (for xlrd)
        xls_copy = make_temp_copy(path, ".xls")
        temps.append(xls_copy)
        xlsx_path = convert_xls_to_xlsx(xls_copy)
        temps.append(xlsx_path)
        return xlsx_path, temps
    else:
        # Just copy with correct extension so openpyxl is happy
        copy_path = make_temp_copy(path, ext)
        temps.append(copy_path)
        return copy_path, temps


def analyze_file(path: str, original_name: str = "") -> dict:
    """Read xlsx structure: sheets, columns, row counts, sample data."""
    temps = []
    try:
        import openpyxl

        work_path, temps = prepare_file(path, original_name)

        wb = openpyxl.load_workbook(work_path, read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(max_row=6, values_only=True))
            if not rows:
                sheets.append({"name": sheet_name, "columns": [], "row_count": 0, "sample": []})
                continue

            headers = [str(c) if c is not None else f"Col{i+1}" for i, c in enumerate(rows[0])]
            row_count = ws.max_row or 0
            sample = []
            for row in rows[1:6]:
                sample.append([str(v) if v is not None else "" for v in row])

            sheets.append({
                "name": sheet_name,
                "columns": headers,
                "row_count": max(0, row_count - 1),
                "sample": sample
            })
        wb.close()
        return {"ok": True, "sheets": sheets}
    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        for t in temps:
            try:
                if os.path.exists(t):
                    os.unlink(t)
            except Exception:
                pass


def execute_code(input_path: str, code: str, output_path: str, original_name: str = "") -> dict:
    """
    Execute AI-generated openpyxl code in a restricted namespace.
    """
    temps = []
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter, column_index_from_string
        import re

        work_path, temps = prepare_file(input_path, original_name)

        wb = openpyxl.load_workbook(work_path)

        from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
        from openpyxl.chart.series import DataPoint
        import datetime
        import math
        from collections import defaultdict, Counter

        safe_globals = {
            "__builtins__": {
                "range": range, "len": len, "print": print,
                "str": str, "int": int, "float": float, "bool": bool,
                "list": list, "dict": dict, "tuple": tuple, "set": set,
                "sorted": sorted, "reversed": reversed,
                "enumerate": enumerate, "zip": zip, "map": map, "filter": filter,
                "min": min, "max": max, "sum": sum, "abs": abs, "round": round,
                "isinstance": isinstance, "type": type, "hasattr": hasattr,
                "getattr": getattr, "setattr": setattr,
                "any": any, "all": all, "next": next, "iter": iter,
                "True": True, "False": False, "None": None,
                "Exception": Exception, "ValueError": ValueError, "KeyError": KeyError,
                "IndexError": IndexError, "TypeError": TypeError,
            },
            "openpyxl": openpyxl,
            "PatternFill": PatternFill,
            "Font": Font,
            "Alignment": Alignment,
            "Border": Border,
            "Side": Side,
            "get_column_letter": get_column_letter,
            "column_index_from_string": column_index_from_string,
            "re": re,
            "datetime": datetime.datetime,  # datetime IS the class, not the module
            "datetime_module": datetime,      # use datetime_module.now() if needed
            # datetime shortcuts — use these directly without datetime. prefix
            "date": datetime.date,
            "time": datetime.time,
            "timedelta": datetime.timedelta,
            "math": math,
            "defaultdict": defaultdict,
            "Counter": Counter,
            # Chart classes
            "BarChart": BarChart,
            "LineChart": LineChart,
            "PieChart": PieChart,
            "Reference": Reference,
            "Series": Series,
            "DataPoint": DataPoint,
            "wb": wb,
            "changes": [],
        }

        exec(code, safe_globals)

        wb.save(output_path)

        changes = safe_globals.get("changes", [])
        return {"ok": True, "changes": changes}

    except Exception as e:
        return {"ok": False, "error": str(e), "traceback": traceback.format_exc()}
    finally:
        for t in temps:
            try:
                if os.path.exists(t):
                    os.unlink(t)
            except Exception:
                pass


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""

    if cmd == "analyze":
        path = sys.argv[2]
        original_name = sys.argv[3] if len(sys.argv) > 3 else ""
        result = analyze_file(path, original_name)
        print(json.dumps(result, ensure_ascii=False))

    elif cmd == "execute":
        input_path = sys.argv[2]
        output_path = sys.argv[3]
        code_file = sys.argv[4]
        original_name = sys.argv[5] if len(sys.argv) > 5 else ""
        with open(code_file, "r", encoding="utf-8") as f:
            code = f.read()
        result = execute_code(input_path, code, output_path, original_name)
        print(json.dumps(result, ensure_ascii=False))

    else:
        print(json.dumps({"ok": False, "error": f"Unknown command: {cmd}"}))
        sys.exit(1)
